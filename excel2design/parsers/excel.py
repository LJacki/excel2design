"""Excel workbook parser per SPEC §2 and §3.

Reads a .xlsx file in the canonical two-section layout and returns a list of
Module objects (one per sheet, except for the @parameters concept which is
deprecated in v0.3 — each module sheet now carries its own parameters).

Public entry points:
  - parse_workbook(path) -> list[Module]
  - get_module(modules, name) -> Module  (raises ModuleNotFoundError)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel2design.core.exceptions import (
    DuplicatePortError,
    ExcelParseError,
    FormulaCellError,
    HeaderMismatchError,
    MarkerMissingError,
    MergedCellError,
    ModuleNotFoundError,
    PortValidationError,
)
from excel2design.core.models import (
    Direction,
    Module,
    Parameter,
    ParamType,
    Port,
    ResetType,
    SignalType,
    Define,
)
from excel2design.parsers.default import DefaultValueError, parse_default
from excel2design.parsers.width import PortWidth, parse_width
from excel2design.utils.cell import cell_to_str, is_blank
from excel2design.utils.identifier import check_identifier


# Markers and canonical column orders
MARKER_PARAMETERS = "# === PARAMETERS ==="
MARKER_PORTS = "# === PORTS ==="
MARKER_DEFINES = "# === DEFINES ==="
DEFINES_SHEET = "@defines"

PARAM_HEADER = ["name", "value", "width", "param_type", "comment"]
PORT_HEADER = [
    "name", "direction", "width", "type", "default",
    "clock", "reset_type", "signed", "interface", "comment",
]


# ---- Public API ------------------------------------------------------------

def parse_workbook(path: Path | str) -> list[Module]:
    """Parse a .xlsx file. Returns one Module per non-empty sheet.
    Skips @defines sheet.

    Raises:
        ExcelParseError: on any structural problem.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    wb = load_workbook(path, data_only=True)
    modules: list[Module] = []
    for sheet_name in wb.sheetnames:
        if sheet_name == DEFINES_SHEET:
            continue
        ws = wb[sheet_name]
        module = _parse_sheet(ws, source_file=path)
        if module is not None:
            modules.append(module)
    return modules


def parse_defines(path: Path | str) -> list[Define]:
    """Parse the @defines sheet only. Returns empty list if no such sheet."""
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    if DEFINES_SHEET in wb.sheetnames:
        return _parse_defines_sheet(wb[DEFINES_SHEET])
    return []


def get_module(modules: list[Module], name: str) -> Module:
    for m in modules:
        if m.name == name:
            return m
    raise ModuleNotFoundError(name, [m.name for m in modules])


# ---- Sheet-level parsing ---------------------------------------------------

def _parse_sheet(ws: Worksheet, source_file: Path) -> Optional[Module]:
    """Parse one sheet into a Module. Returns None if the sheet is empty."""
    # Detect merged cells anywhere — v0.3 doesn't support them
    if ws.merged_cells.ranges:
        from excel2design.core.exceptions import MergedCellError
        ranges = ", ".join(str(r) for r in ws.merged_cells.ranges)
        raise MergedCellError(ws.title, ranges)

    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return None

    # Find markers and section boundaries
    param_marker_row = _find_marker(ws, rows, MARKER_PARAMETERS)
    port_marker_row = _find_marker(ws, rows, MARKER_PORTS)

    # If neither marker present, treat sheet as empty (skip)
    if param_marker_row is None and port_marker_row is None:
        return None
    if param_marker_row is None:
        raise MarkerMissingError(ws.title, MARKER_PARAMETERS)
    if port_marker_row is None:
        raise MarkerMissingError(ws.title, MARKER_PORTS)

    # Validate sheet name as Verilog identifier (used as module name).
    # For dotted names (hierarchy), validate only the last segment.
    raw_name = ws.title
    if "." in raw_name:
        module_name = raw_name.rsplit(".", 1)[-1]
    else:
        module_name = raw_name
    check_identifier(module_name, "module (sheet name)", sheet=ws.title, row=0, col=0)

    # ---- Parameter section: header is param_marker_row + 1, data after ----
    param_header_row = param_marker_row + 1
    _validate_header(ws, param_header_row, PARAM_HEADER, "PARAMETERS")
    # Data rows: skip blanks/comments, until we hit the port marker or sheet end
    data_start = param_header_row + 1
    data_end = port_marker_row  # exclusive
    parameters = _parse_parameters(ws, data_start, data_end, source_sheet=ws.title)

    # ---- Port section: header is port_marker_row + 1, data after ------
    port_header_row = port_marker_row + 1
    _validate_header(ws, port_header_row, PORT_HEADER, "PORTS")
    ports = _parse_ports(ws, port_header_row + 1, source_sheet=ws.title, known_params=set(p.name for p in parameters))

    # ---- Validate uniqueness ---------------------------------------------
    _validate_unique_ports(ports, source_sheet=ws.title)

    return Module(
        name=module_name,  # last segment for dotted names
        ports=ports,
        parameters=parameters,
        source_file=source_file,
        source_sheet=ws.title,
    )


def _find_marker(ws: Worksheet, rows: list[Any], marker: str) -> Optional[int]:
    """Return the 1-based row number where `marker` appears in column A, or None."""
    for idx, row in enumerate(rows, start=1):
        if not row:
            continue
        first = row[0]
        if first.value is None:
            continue
        if isinstance(first.value, str) and first.value.strip() == marker:
            return idx
    return None


def _validate_header(ws: Worksheet, row_num: int, expected: list[str], section: str) -> None:
    """Verify the header row matches the expected column order (case-insensitive)."""
    row = ws[row_num]
    actual = [cell_to_str(c) for c in row[:len(expected)]]
    if [a.lower() for a in actual] != [e.lower() for e in expected]:
        raise HeaderMismatchError(ws.title, row_num, expected, actual)


# ---- Parameter section -----------------------------------------------------

def _parse_parameters(
    ws: Worksheet, start: int, end: int, source_sheet: str
) -> list[Parameter]:
    """Parse parameter rows from row `start` to `end` (exclusive)."""
    params: list[Parameter] = []
    for r in range(start, end + 1):
        row = ws[r]
        if not row:
            continue
        # Skip fully-blank rows
        if all(is_blank(c) for c in row[:5]):
            continue
        # Skip comment rows (first cell starts with '#')
        first_val = cell_to_str(row[0])
        if first_val.startswith("#"):
            continue

        # Required: name, value
        name = cell_to_str(row[0])
        value = cell_to_str(row[1])
        if not name:
            # Skip rows that have no name and no value (likely stray)
            if not value:
                continue
            raise ExcelParseError(
                "parameter 缺 name",
                sheet=source_sheet, row=r, col=1,
                suggestion="name 列必须填写",
            )
        if not value:
            raise ExcelParseError(
                f"parameter '{name}' 缺 value",
                sheet=source_sheet, row=r, col=2,
                suggestion="value 列必须填写",
            )

        check_identifier(name, "parameter", sheet=source_sheet, row=r, col=1)

        width = cell_to_str(row[2]) or None
        param_type = ParamType.parse(cell_to_str(row[3]))
        comment = cell_to_str(row[4]) or None

        params.append(Parameter(
            name=name,
            value=value,
            width=width,
            param_type=param_type,
            comment=comment,
        ))
    return params


# ---- Port section ----------------------------------------------------------

def _parse_ports(
    ws: Worksheet, start: int, source_sheet: str, known_params: set[str]
) -> list[Port]:
    """Parse port rows starting at row `start` (header was at start-1)."""
    ports: list[Port] = []
    # Find the last row that has any data
    last_row = ws.max_row
    for r in range(start, last_row + 1):
        row = ws[r]
        if not row:
            continue
        if all(is_blank(c) for c in row[:10]):
            continue
        first_val = cell_to_str(row[0])
        if first_val.startswith("#"):
            continue

        try:
            port = _parse_one_port(row, r, source_sheet, known_params)
        except (PortValidationError, ExcelParseError):
            raise
        if port is not None:
            ports.append(port)
    return ports


def _parse_one_port(
    row: tuple[Any, ...], row_num: int, source_sheet: str, known_params: set[str]
) -> Optional[Port]:
    name = cell_to_str(row[0])
    direction_raw = cell_to_str(row[1])
    width_raw = cell_to_str(row[2])
    type_raw = cell_to_str(row[3])
    default_raw = cell_to_str(row[4])
    clock_raw = cell_to_str(row[5])
    reset_raw = cell_to_str(row[6])
    signed_raw = cell_to_str(row[7])
    interface_raw = cell_to_str(row[8])
    comment = cell_to_str(row[9])

    if not name:
        # No name = skip (treated as stray)
        return None

    check_identifier(name, "port", sheet=source_sheet, row=row_num, col=1)

    direction = Direction.parse(direction_raw)
    width = parse_width(width_raw, known_params, sheet=source_sheet, row=row_num, col=3)
    signal_type = SignalType.parse(type_raw, direction)

    default: Optional[str] = None
    if default_raw:
        try:
            default = parse_default(default_raw)
        except DefaultValueError as e:
            raise PortValidationError(
                str(e),
                sheet=source_sheet, row=row_num, col=5,
                suggestion="default 必须是 Verilog 字面量（如 1'b0, 8'hFF）或 replication（如 {N{1'b0}}）",
            ) from e

    clock = clock_raw or None
    reset_type = ResetType.parse(reset_raw)

    # signed: accept 0/1, "true"/"false", "yes"/"no"
    signed = _truthy(signed_raw)
    is_interface = _truthy(interface_raw)

    return Port(
        name=name,
        direction=direction,
        width=width,
        type=signal_type,
        default=default,
        clock=clock,
        reset_type=reset_type,
        signed=signed,
        is_interface=is_interface,
        comment=comment or None,
    )


def _truthy(raw: str) -> bool:
    """Parse a 'truthy' cell value: 1, true, yes (case-insensitive)."""
    return raw.strip().lower() in {"1", "true", "yes", "y", "t"}


def _validate_unique_ports(ports: list[Port], source_sheet: str) -> None:
    rows_by_name: dict[str, list[int]] = {}
    for i, p in enumerate(ports, start=1):
        rows_by_name.setdefault(p.name, []).append(i)
    for name, rows in rows_by_name.items():
        if len(rows) > 1:
            raise DuplicatePortError(source_sheet, name, rows)


# ---- v0.5: @defines sheet parsing ------------------------------------------

def _parse_defines_sheet(ws: Worksheet) -> list[Define]:
    """Parse the @defines sheet. Returns list of Define entries."""
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return []

    # Find the DEFINES marker
    marker_row = _find_marker(ws, rows, MARKER_DEFINES)
    if marker_row is None:
        return []

    # Header row follows the marker
    header_row = marker_row + 1
    if header_row > len(rows):
        return []
    header_cells = rows[header_row - 1]  # rows is 0-indexed
    header = [str(c.value).strip().lower() if c.value else "" for c in header_cells[:3]]
    expected = ["name", "value", "comment"]
    if header != expected:
        raise HeaderMismatchError(ws.title, header_row, expected, header)

    defines: list[Define] = []
    for r in range(header_row + 1, len(rows) + 1):
        cells = rows[r - 1]  # rows is 0-indexed
        name = str(cells[0].value).strip() if cells[0].value else ""
        if not name or name.startswith("#"):
            continue
        raw_value = cells[1].value
        value = str(raw_value).strip() if raw_value is not None else ""
        comment = str(cells[2].value).strip() if len(cells) > 2 and cells[2].value else None
        defines.append(Define(name=name, value=value, comment=comment))

    return defines
