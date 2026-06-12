#!/usr/bin/env python3
"""Generate tests/fixtures/interface_ports.xlsx (Phase 13 fixture).

Run from repo root:
    python tools/gen_interface_fixture.py

The fixture contains one parameter-less module ``bus_consumer`` with a mix
of plain ports and interface-grouped ports (marked with ``interface=1``).

Used by tests/generators/test_verilog.py::test_interface_*
and tests/generators/test_diagram_width_none.py::test_interface_*
"""

from __future__ import annotations

import sys
from pathlib import Path

# Allow running as a script without installing
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = REPO_ROOT / "tests" / "fixtures" / "interface_ports.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MARKER_FONT = Font(bold=True, color="333333", italic=True)
MARKER_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER


def style_marker(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = MARKER_FONT
        c.fill = MARKER_FILL


def write_sheet(ws, parameters, ports) -> None:
    PARAM_COLS = 5
    PORT_COLS = 10
    # Parameters block
    ws.cell(row=1, column=1, value="# === PARAMETERS ===")
    style_marker(ws, 1, PARAM_COLS)
    for c, h in enumerate(["name", "value", "width", "param_type", "comment"], 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, PARAM_COLS)
    for r, p in enumerate(parameters, start=3):
        for c, v in enumerate(p, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    next_row = 3 + len(parameters) + 1
    # Ports block
    ws.cell(row=next_row, column=1, value="# === PORTS ===")
    style_marker(ws, next_row, PORT_COLS)
    next_row += 1
    for c, h in enumerate(
        [
            "name", "direction", "width", "type", "default",
            "clock", "reset_type", "signed", "interface", "comment",
        ],
        1,
    ):
        ws.cell(row=next_row, column=c, value=h)
    style_header(ws, next_row, PORT_COLS)
    next_row += 1
    for r, p in enumerate(ports, start=next_row):
        for c, v in enumerate(p, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    widths = [18, 12, 14, 10, 22, 8, 12, 8, 10, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# bus_consumer: 8 ports, 5 of which are interface-grouped (axi_in bus).
# Pattern matches typical SV interface usage: input clk/rst_n + a few
# interface-tagged inputs/outputs that would be split into multiple
# physical wires at synthesis time (v0 仅记录，不做 modport 展开).
PARAMS: list[tuple] = []
PORTS: list[tuple] = [
    # name          direction width  type  default      clock reset_type signed interface comment
    ("clk",          "input",  1, "wire", "",          "",    "",         0, 0, "系统时钟"),
    ("rst_n",        "input",  1, "wire", "",          "",    "",         0, 0, "异步低有效复位"),
    # ---- interface axi_in members ----
    ("bus_data",     "input",  8, "wire", "",          "",    "",         0, 1, "interface axi_in"),
    ("bus_valid",    "input",  1, "wire", "",          "",    "",         0, 1, "interface axi_in"),
    ("bus_ready",    "output", 1, "reg",  "1'b0",      "clk", "async",    0, 1, "interface axi_in"),
    # ---- plain (non-interface) ports ----
    ("irq_out",      "output", 1, "reg",  "1'b0",      "clk", "async",    0, 0, "普通输出"),
    ("status_reg",   "output", 4, "reg",  "4'h0",      "clk", "sync",     0, 0, "普通状态"),
]


def main() -> int:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "bus_consumer"
    write_sheet(ws, PARAMS, PORTS)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
