#!/usr/bin/env python3
"""Generate the canonical sample Excel used for testing and documentation.

Run from repo root:
    python tools/gen_sample.py

Output: examples/sample_module_uart_rx.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

# Allow running as a script without installing the package
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = REPO_ROOT / "examples" / "sample_module_uart_rx.xlsx"

# ---- Styling helpers ------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MARKER_FONT = Font(bold=True, color="333333", italic=True)
MARKER_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def style_marker(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = MARKER_FONT
        cell.fill = MARKER_FILL


def write_sheet(ws, parameters: list[tuple], ports: list[tuple]) -> None:
    """Write a module sheet in the canonical two-section layout.

    Parameters and ports are lists of tuples matching the column order in SPEC §2.4/2.5.
    """
    PARAM_COLS = 5   # name | value | width | param_type | comment
    PORT_COLS = 10   # name | direction | width | type | default | clock | reset_type | signed | interface | comment

    # Row 1: parameter marker
    ws.cell(row=1, column=1, value="# === PARAMETERS ===")
    style_marker(ws, 1, PARAM_COLS)

    # Row 2: parameter header
    param_header = ["name", "value", "width", "param_type", "comment"]
    for c, h in enumerate(param_header, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, PARAM_COLS)

    # Rows 3..: parameter rows
    for r, row_data in enumerate(parameters, start=3):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # Find the first empty row after parameters
    next_row = 3 + len(parameters)

    # Section separator (optional blank row)
    if parameters:
        next_row += 1

    # Marker for ports section
    ws.cell(row=next_row, column=1, value="# === PORTS ===")
    style_marker(ws, next_row, PORT_COLS)
    next_row += 1

    # Port header
    port_header = [
        "name", "direction", "width", "type", "default",
        "clock", "reset_type", "signed", "interface", "comment",
    ]
    for c, h in enumerate(port_header, 1):
        ws.cell(row=next_row, column=c, value=h)
    style_header(ws, next_row, PORT_COLS)
    next_row += 1

    # Port rows
    for r, row_data in enumerate(ports, start=next_row):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # Set sensible column widths
    widths = [18, 12, 14, 10, 22, 8, 12, 8, 10, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Module definitions ---------------------------------------------------

UART_RX_PARAMETERS = [
    # (name,        value, width, param_type, comment)
    ("DATA_WIDTH",   8,   32, "parameter", "数据位宽"),
    ("FIFO_DEPTH",  16,   32, "parameter", "FIFO 深度"),
    ("CLK_FREQ_MHZ", 100, 32, "parameter", "时钟频率(MHz)"),
]

UART_RX_PORTS = [
    # (name,       direction, width,       type, default,                    clock, reset_type, signed, interface, comment)
    ("clk",        "input",  1,            "wire", "",                         "",    "",          0,      0, "系统时钟"),
    ("rst_n",      "input",  1,            "wire", "",                         "",    "",          0,      0, "异步低有效复位"),
    ("rx_pad",     "input",  1,            "wire", "",                         "",    "",          0,      0, "串行输入"),
    ("baud_tick",  "input",  1,            "wire", "",                         "",    "",          0,      0, "波特率 tick"),
    ("rx_data",    "output", "DATA_WIDTH", "reg",  "{DATA_WIDTH{1'b0}}",      "clk", "async",     0,      0, "接收数据"),
    ("rx_valid",   "output", 1,            "reg",  "1'b0",                     "clk", "async",     0,      0, "接收有效"),
    ("fifo_full",  "output", 1,            "reg",  "1'b0",                     "clk", "async",     0,      0, "FIFO 满"),
    ("fifo_data",  "output", "DATA_WIDTH", "reg",  "{DATA_WIDTH{1'b0}}",      "clk", "async",     1,      0, "FIFO 数据 (signed)"),
]


def main() -> int:
    wb = Workbook()
    # Rename the default sheet to our first module
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "uart_rx"
    write_sheet(ws, UART_RX_PARAMETERS, UART_RX_PORTS)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
