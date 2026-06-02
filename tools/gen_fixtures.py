#!/usr/bin/env python3
"""Generate golden test fixtures (4 .xlsx files) used by tests/fixtures/.

Run from repo root:
    python tools/gen_fixtures.py

Outputs:
    tests/fixtures/uart_rx.xlsx         (canonical, 8 ports, parameterized)
    tests/fixtures/axi_crossbar.xlsx    (30 ports, mixed reset, signed, inout)
    tests/fixtures/multi_clock.xlsx     (3 clocks, multi-domain)
    tests/fixtures/empty_ports.xlsx     (parameters only, no ports)
"""

from __future__ import annotations

import sys
from pathlib import Path

# Allow running as a script without installing
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = REPO_ROOT / "tests" / "fixtures"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MARKER_FONT = Font(bold=True, color="333333", italic=True)
MARKER_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER


def style_marker(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = MARKER_FONT
        c.fill = MARKER_FILL


def write_sheet(ws, parameters, ports) -> None:
    PARAM_COLS = 5
    PORT_COLS = 10
    ws.cell(row=1, column=1, value="# === PARAMETERS ===")
    style_marker(ws, 1, PARAM_COLS)
    for c, h in enumerate(["name", "value", "width", "param_type", "comment"], 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, PARAM_COLS)
    for r, p in enumerate(parameters, start=3):
        for c, v in enumerate(p, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    next_row = 3 + len(parameters) + 1
    ws.cell(row=next_row, column=1, value="# === PORTS ===")
    style_marker(ws, next_row, PORT_COLS)
    next_row += 1
    for c, h in enumerate([
        "name", "direction", "width", "type", "default",
        "clock", "reset_type", "signed", "interface", "comment"
    ], 1):
        ws.cell(row=next_row, column=c, value=h)
    style_header(ws, next_row, PORT_COLS)
    next_row += 1
    for r, p in enumerate(ports, start=next_row):
        for c, v in enumerate(p, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    widths = [18, 12, 14, 10, 22, 8, 12, 8, 10, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Fixture 1: uart_rx (canonical, copy of sample) ----------------------

UART_PARAMS = [
    ("DATA_WIDTH",   8,   32, "parameter", "数据位宽"),
    ("FIFO_DEPTH",  16,   32, "parameter", "FIFO 深度"),
    ("CLK_FREQ_MHZ", 100, 32, "parameter", "时钟频率(MHz)"),
]
UART_PORTS = [
    ("clk",       "input",  1,            "wire", "",                     "",    "",         0, 0, "系统时钟"),
    ("rst_n",     "input",  1,            "wire", "",                     "",    "",         0, 0, "异步低有效复位"),
    ("rx_pad",    "input",  1,            "wire", "",                     "",    "",         0, 0, "串行输入"),
    ("baud_tick", "input",  1,            "wire", "",                     "",    "",         0, 0, "波特率 tick"),
    ("rx_data",   "output", "DATA_WIDTH", "reg",  "{DATA_WIDTH{1'b0}}",  "clk", "async",    0, 0, "接收数据"),
    ("rx_valid",  "output", 1,            "reg",  "1'b0",                 "clk", "async",    0, 0, "接收有效"),
    ("fifo_full", "output", 1,            "reg",  "1'b0",                 "clk", "async",    0, 0, "FIFO 满"),
    ("fifo_data", "output", "DATA_WIDTH", "reg",  "{DATA_WIDTH{1'b0}}",  "clk", "async",    1, 0, "FIFO 数据"),
]


# ---- Fixture 2: axi_crossbar (30 ports, mixed, inout) --------------------

AXI_PARAMS = [
    ("ADDR_WIDTH", 32, 32, "parameter", "地址位宽"),
    ("DATA_WIDTH", 64, 32, "parameter", "数据位宽"),
    ("ID_WIDTH",    4, 32, "parameter", "ID 位宽"),
    ("NUM_SLAVES",  4, 32, "parameter", "从机数量"),
]
# Inputs (10) + outputs (15) + inouts (5) = 30
AXI_INPUTS = [
    ("aclk",          "input",  1,                  "wire", "", "", "", 0, 0, "AXI 时钟"),
    ("aresetn",       "input",  1,                  "wire", "", "", "", 0, 0, "AXI 复位"),
    ("s_axi_awid",    "input",  "ID_WIDTH",         "wire", "", "", "", 0, 0, ""),
    ("s_axi_awaddr",  "input",  "ADDR_WIDTH",       "wire", "", "", "", 0, 0, ""),
    ("s_axi_awlen",   "input",  8,                  "wire", "", "", "", 0, 0, ""),
    ("s_axi_awsize",  "input",  3,                  "wire", "", "", "", 0, 0, ""),
    ("s_axi_awburst", "input",  2,                  "wire", "", "", "", 0, 0, ""),
    ("s_axi_awvalid", "input",  1,                  "wire", "", "", "", 0, 0, ""),
    ("s_axi_wdata",   "input",  "DATA_WIDTH",       "wire", "", "", "", 0, 0, ""),
    ("s_axi_wstrb",   "input",  "DATA_WIDTH/8",     "wire", "", "", "", 0, 0, ""),
]
AXI_OUTPUTS = [
    ("s_axi_awready", "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("s_axi_wready",  "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("s_axi_bvalid",  "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("s_axi_bresp",   "output", 2, "reg", "2'b00", "aclk", "async", 0, 0, ""),
    ("s_axi_arready", "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("s_axi_rvalid",  "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("s_axi_rdata",   "output", "DATA_WIDTH", "reg", "{DATA_WIDTH{1'b0}}", "aclk", "async", 1, 0, ""),
    ("s_axi_rresp",   "output", 2, "reg", "2'b00", "aclk", "async", 0, 0, ""),
    ("s_axi_rlast",   "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("m_axi_awid",    "output", "ID_WIDTH",   "reg", "{ID_WIDTH{1'b0}}",   "aclk", "async", 0, 0, ""),
    ("m_axi_awaddr",  "output", "ADDR_WIDTH", "reg", "{ADDR_WIDTH{1'b0}}", "aclk", "async", 0, 0, ""),
    ("m_axi_awvalid", "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("m_axi_wdata",   "output", "DATA_WIDTH", "reg", "{DATA_WIDTH{1'b0}}", "aclk", "async", 0, 0, ""),
    ("m_axi_wvalid",  "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
    ("m_axi_bready",  "output", 1, "reg", "1'b0", "aclk", "async", 0, 0, ""),
]
AXI_INOUTS = [
    ("s_axi_awuser", "inout", 1, "wire", "", "", "", 0, 0, "user signal"),
    ("s_axi_aruser", "inout", 1, "wire", "", "", "", 0, 0, "user signal"),
    ("s_axi_buser",  "inout", 1, "wire", "", "", "", 0, 0, "user signal"),
    ("s_axi_ruser",  "inout", 1, "wire", "", "", "", 0, 0, "user signal"),
    ("debug",        "inout", 8, "wire", "", "", "", 0, 0, "debug bus"),
]


# ---- Fixture 3: multi_clock (3 clock domains) -----------------------------

MULTI_PARAMS = [
    ("WIDTH", 16, 32, "parameter", "数据位宽"),
]
MULTI_PORTS = [
    # clk_a domain
    ("clk_a",    "input",  1,        "wire", "", "",                  "",  0, 0, "domain A clock"),
    ("rst_a_n",  "input",  1,        "wire", "", "",                  "",  0, 0, "domain A reset"),
    ("data_a",   "output", "WIDTH",  "reg",  "{WIDTH{1'b0}}",        "clk_a", "async", 0, 0, "data a"),
    ("valid_a",  "output", 1,        "reg",  "1'b0",                 "clk_a", "async", 0, 0, "valid a"),
    # clk_b domain
    ("clk_b",    "input",  1,        "wire", "", "",                  "",  0, 0, "domain B clock"),
    ("rst_b_n",  "input",  1,        "wire", "", "",                  "",  0, 0, "domain B reset"),
    ("data_b",   "output", "WIDTH",  "reg",  "{WIDTH{1'b0}}",        "clk_b", "async", 0, 0, "data b"),
    # clk_c domain (sync reset, no default)
    ("clk_c",    "input",  1,        "wire", "", "",                  "",  0, 0, "domain C clock"),
    ("flag_c",   "output", 1,        "reg",  "1'b0",                 "clk_c", "sync",  0, 0, "flag c (sync reset)"),
    # cross-domain
    ("bridge_in",  "input",  "WIDTH", "wire", "", "",                 "",  0, 0, "from a to b"),
    ("bridge_out", "output", "WIDTH", "reg",  "{WIDTH{1'b0}}",        "clk_b", "async", 0, 0, "to a from b"),
]


# ---- Fixture 4: empty_ports (parameters only) -----------------------------

EMPTY_PARAMS = [
    ("VERSION",   1,   32, "parameter",  "版本号"),
    ("BUILD_DATE", "20260601", 32, "parameter", "构建日期"),
]
EMPTY_PORTS: list = []


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    fixtures = [
        ("uart_rx.xlsx",      "uart_rx",      UART_PARAMS, UART_PORTS),
        ("axi_crossbar.xlsx", "axi_crossbar", AXI_PARAMS,   AXI_INPUTS + AXI_OUTPUTS + AXI_INOUTS),
        ("multi_clock.xlsx",  "multi_clock",  MULTI_PARAMS, MULTI_PORTS),
        ("empty_ports.xlsx",  "empty_ports",  EMPTY_PARAMS, EMPTY_PORTS),
    ]

    for filename, sheet_name, params, ports in fixtures:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name
        write_sheet(ws, params, ports)
        path = OUT_DIR / filename
        wb.save(path)
        print(f"Wrote {path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
