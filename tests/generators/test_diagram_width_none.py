"""P0-3 fix: 3 diagram generators must tolerate width=None (default 1-bit).

Background:
  v0.3.2 fixed PortWidth.to_verilog() for wrapper, but the 3 diagram
  generators (HTML/SVG/Excalidraw) still had `assert p.width.msb is not None`.
  This caused AssertionError when the user ran `diagram` or `all` on a
  module with a blank width cell.
"""

from __future__ import annotations

import json
import xml.etree.ElementTree as ET
from pathlib import Path

import pytest
from openpyxl import Workbook

from excel2design import parse_workbook, get_module
from excel2design.generators.diagram_html import generate_html
from excel2design.generators.diagram_svg import generate_svg
from excel2design.generators.diagram_excalidraw import generate_excalidraw

FIXTURE_DIR = Path(__file__).resolve().parents[2] / "tests" / "fixtures"


@pytest.fixture
def module_with_blank_width(tmp_path: Path):
    """Create an Excel module with a blank width cell, then return the parsed Module."""
    wb = Workbook()
    ws = wb.active
    ws.title = "no_width"
    # Parameters
    ws.cell(1, 1, "# === PARAMETERS ===")
    for c, h in enumerate(["name", "value", "width", "param_type", "comment"], 1):
        ws.cell(2, c, h)
    ws.cell(3, 1, "W")
    ws.cell(3, 2, 8)
    # Ports
    ws.cell(4, 1, "# === PORTS ===")
    for c, h in enumerate(
        [
            "name", "direction", "width", "type", "default",
            "clock", "reset_type", "signed", "interface", "comment",
        ],
        1,
    ):
        ws.cell(5, c, h)
    # One port with blank width, one with explicit '1' (sanity check)
    ws.cell(6, 1, "data_blank")
    ws.cell(6, 2, "output")
    ws.cell(6, 3, None)  # BLANK width
    ws.cell(6, 4, "reg")
    ws.cell(6, 5, "1'b0")
    ws.cell(6, 6, "clk")
    ws.cell(6, 7, "async")
    ws.cell(7, 1, "data_explicit")
    ws.cell(7, 2, "output")
    ws.cell(7, 3, 1)  # explicit 1-bit
    ws.cell(7, 4, "reg")
    ws.cell(7, 5, "1'b0")
    ws.cell(7, 6, "clk")
    ws.cell(7, 7, "async")
    p = tmp_path / "no_width.xlsx"
    wb.save(p)
    modules = parse_workbook(p)
    return modules[0]


# ---- HTML -----------------------------------------------------------------

def test_html_handles_blank_width(module_with_blank_width) -> None:
    """P0-3: HTML generator must not AssertionError on width=None."""
    html = generate_html(module_with_blank_width)
    # Port should be rendered with empty width badge
    assert "data_blank" in html


def test_html_blank_width_does_not_leak_none(module_with_blank_width) -> None:
    html = generate_html(module_with_blank_width)
    assert "None" not in html


# ---- SVG ------------------------------------------------------------------

def test_svg_handles_blank_width(module_with_blank_width) -> None:
    """P0-3: SVG generator must not AssertionError on width=None."""
    svg = generate_svg(module_with_blank_width)
    # Must be valid XML
    root = ET.fromstring(svg)
    assert root.tag.endswith("svg")


def test_svg_blank_width_does_not_leak_none(module_with_blank_width) -> None:
    svg = generate_svg(module_with_blank_width)
    assert "None" not in svg


# ---- Excalidraw -----------------------------------------------------------

def test_excalidraw_handles_blank_width(module_with_blank_width) -> None:
    """P0-3: Excalidraw generator must not AssertionError on width=None."""
    data = json.loads(generate_excalidraw(module_with_blank_width))
    assert data["type"] == "excalidraw"
    # v0.4: port labels are on arrow elements
    arrows = [el for el in data["elements"] if el["type"] == "arrow"]
    labels = [el["text"] for el in arrows]
    assert "data_blank" in labels


def test_excalidraw_blank_width_does_not_leak_none(module_with_blank_width) -> None:
    data = json.loads(generate_excalidraw(module_with_blank_width))
    raw = generate_excalidraw(module_with_blank_width)
    assert "None" not in raw
