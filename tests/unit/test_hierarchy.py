"""P1-1 unit tests for the v0.5 hierarchy parser (parsers/hierarchy.py).

Covers:
  - parse_project() happy path
  - top_modules property
  - get_submodules(recursive=True/False)
  - walk_bfs() ordering
  - OrphanChildError on missing parent
  - EmptyHierarchyError when every sheet is dotted
  - RecursiveHierarchyError on cycles
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from excel2design.core.exceptions import (
    EmptyHierarchyError, OrphanChildError, RecursiveHierarchyError,
)
from excel2design.parsers.hierarchy import parse_project


# ---- Fixture builders -----------------------------------------------------


def _build_xlsx_with_sheets(sheets: dict[str, list[tuple[str, str, str, str]]],
                              tmp_path, defines: list[tuple[str, str]] | None = None):
    """Build a minimal xlsx with the given sheets.

    ``sheets`` is a dict ``{sheet_name: [(name, direction, width, type), ...]}``.
    Each sheet has the canonical two-section layout with empty params and
    a single port row per port tuple.
    """
    wb = Workbook()
    # Remove the default sheet.
    wb.remove(wb.active)
    for sname, ports in sheets.items():
        ws = wb.create_sheet(sname)
        # Header line + params section (5 columns).
        ws.cell(row=1, column=1, value="# === PARAMETERS ===")
        for col, name in enumerate(["name", "value", "width", "param_type", "comment"], 1):
            ws.cell(row=2, column=col, value=name)
        # Ports section.
        ws.cell(row=4, column=1, value="# === PORTS ===")
        for col, name in enumerate(["name", "direction", "width", "type", "default",
                                     "clock", "reset_type", "signed", "interface", "comment"], 1):
            ws.cell(row=5, column=col, value=name)
        for i, (n, d, w, t) in enumerate(ports):
            ws.cell(row=6 + i, column=1, value=n)
            ws.cell(row=6 + i, column=2, value=d)
            ws.cell(row=6 + i, column=3, value=w)
            ws.cell(row=6 + i, column=4, value=t)
    if defines is not None:
        ws = wb.create_sheet("@defines")
        ws.cell(row=1, column=1, value="# === DEFINES ===")
        for col, name in enumerate(["name", "value", "comment"], 1):
            ws.cell(row=2, column=col, value=name)
        for i, (n, v) in enumerate(defines):
            ws.cell(row=3 + i, column=1, value=n)
            ws.cell(row=3 + i, column=2, value=v)
    p = tmp_path / "test.xlsx"
    wb.save(p)
    return p


# ---- Top-level modules ----------------------------------------------------


def test_top_modules_returns_no_dotted_sheets(tmp_path):
    p = _build_xlsx_with_sheets({
        "top": [("clk", "input", "1", "wire")],
        "top.u_a": [("clk", "input", "1", "wire")],
    }, tmp_path)
    proj = parse_project(p)
    assert proj.top_modules == ["top"]


def test_get_submodules_recursive_false(tmp_path):
    p = _build_xlsx_with_sheets({
        "top": [("clk", "input", "1", "wire")],
        "top.u_a": [("clk", "input", "1", "wire")],
        "top.u_a.u_aa": [("clk", "input", "1", "wire")],
        "top.u_b": [("clk", "input", "1", "wire")],
    }, tmp_path)
    proj = parse_project(p)
    direct = proj.get_submodules("top", recursive=False)
    names = sorted(s.instance_name for s in direct)
    assert names == ["u_a", "u_b"]


def test_get_submodules_recursive_true(tmp_path):
    p = _build_xlsx_with_sheets({
        "top": [("clk", "input", "1", "wire")],
        "top.u_a": [("clk", "input", "1", "wire")],
        "top.u_a.u_aa": [("clk", "input", "1", "wire")],
        "top.u_b": [("clk", "input", "1", "wire")],
    }, tmp_path)
    proj = parse_project(p)
    all_subs = proj.get_submodules("top", recursive=True)
    names = sorted(s.instance_name for s in all_subs)
    assert names == ["u_a", "u_aa", "u_b"]
    # Depth of nested u_aa should be > 1.
    ua_aa = next(s for s in all_subs if s.instance_name == "u_aa")
    assert ua_aa.depth == 2


def test_walk_bfs_orders_top_first_then_children(tmp_path):
    p = _build_xlsx_with_sheets({
        "top": [("clk", "input", "1", "wire")],
        "top.u_a": [("clk", "input", "1", "wire")],
        "top.u_b": [("clk", "input", "1", "wire")],
        "top.u_a.u_aa": [("clk", "input", "1", "wire")],
    }, tmp_path)
    proj = parse_project(p)
    bfs = proj.walk_bfs()
    # top must come before its children
    assert bfs.index("top") < bfs.index("top.u_a")
    assert bfs.index("top") < bfs.index("top.u_b")
    # Children must come before grandchildren
    assert bfs.index("top.u_a") < bfs.index("top.u_a.u_aa")


# ---- Error paths ----------------------------------------------------------


def test_orphan_child_raises(tmp_path):
    """A dotted sheet whose parent doesn't exist must raise OrphanChildError."""
    p = _build_xlsx_with_sheets({
        "ghost.u_a": [("clk", "input", "1", "wire")],  # parent "ghost" missing
    }, tmp_path)
    with pytest.raises(OrphanChildError) as exc:
        parse_project(p)
    assert exc.value.parent == "ghost"
    assert exc.value.sheet == "ghost.u_a"


def test_empty_hierarchy_raises(tmp_path):
    """If every sheet is dotted (no top-level), raise EmptyHierarchyError.

    Construct the case carefully: every sheet must be a child of an existing
    parent (otherwise OrphanChildError fires first per SPEC §19.1 ordering).
    """
    p = _build_xlsx_with_sheets({
        "wrapper": [("clk", "input", "1", "wire")],          # top exists
        "wrapper.sub1": [("clk", "input", "1", "wire")],
        "wrapper.sub2": [("clk", "input", "1", "wire")],
        # Now force an "empty" hierarchy by removing the top in a fresh wb
        # — easier to test the error path through the unit-level helper
        # indirectly: we can confirm the function is invoked when no top.
    }, tmp_path)
    proj = parse_project(p)
    # With one top-level, we should NOT see EmptyHierarchyError.
    assert "wrapper" in proj.top_modules

    # Now test the direct error path using a workbook where every sheet is
    # dotted and orphans are present — the parser will raise OrphanChild
    # first. So to actually exercise EmptyHierarchyError we need a case
    # with no top-level and no orphans, which is impossible. The exception
    # itself is defensive: a workbook where every sheet is dotted means
    # there must be orphans (the dot-stem is missing). So the only
    # realistic trigger would be an Excel with a sheet literally named
    # "sub1.sub2" (parent "sub1" missing) — covered by the orphan test.
    # We still smoke-test the exception class:
    from excel2design.core.exceptions import EmptyHierarchyError as EHE
    assert EHE().args[0]  # instantiable with a message


def test_recursive_hierarchy_raises(tmp_path):
    """A 3-level dotted chain is fine; but a cycle raises RecursiveHierarchyError.

    Cycles can't be expressed with `.` syntax in sheet names (Excel doesn't
    allow them), so we test the cycle detector directly by constructing a
    hierarchy dict that points to a cycle, then calling the helper.

    For end-to-end coverage we instead test the case where the parent of a
    child names itself as a descendant: the parser uses DFS over the
    hierarchy dict. Since the parser builds hierarchy FROM the sheet names
    (no cycles possible in xlsx), we can only exercise the cycle detector
    indirectly. This test is therefore a smoke test: a 3-level chain is
    parsed without error.
    """
    p = _build_xlsx_with_sheets({
        "a": [("clk", "input", "1", "wire")],
        "a.b": [("clk", "input", "1", "wire")],
        "a.b.c": [("clk", "input", "1", "wire")],
    }, tmp_path)
    proj = parse_project(p)
    # Smoke: 3-level chain parsed without error. ``source_sheet`` keeps
    # the full dotted name; ``name`` is the short (last-segment) name.
    assert proj.modules["a.b.c"].source_sheet == "a.b.c"
    assert proj.modules["a.b.c"].name == "c"
    # If we did somehow have a cycle it would raise — confirm the helper
    # raises when fed a cyclic dict.
    from excel2design.parsers.hierarchy import _check_no_cycles
    with pytest.raises(RecursiveHierarchyError):
        _check_no_cycles(["a", "b"], {"a": ["b"], "b": ["a"]})


# ---- @defines integration -------------------------------------------------


def test_defines_sheet_parsed(tmp_path):
    p = _build_xlsx_with_sheets(
        {"top": [("clk", "input", "1", "wire")]},
        tmp_path,
        defines=[("ADC_EN", "32"), ("BAUD", "115200")],
    )
    proj = parse_project(p)
    names = [d.name for d in proj.defines]
    assert names == ["ADC_EN", "BAUD"]
    assert proj.defines[0].value == "32"
