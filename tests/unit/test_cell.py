"""Tests for utils/cell.py (SPEC §3.5.1)."""

from __future__ import annotations

import datetime

import pytest
from openpyxl import Workbook

from excel2design.core.exceptions import FormulaCellError, UnsupportedCellTypeError
from excel2design.utils.cell import cell_to_str, is_blank


@pytest.fixture
def ws():
    wb = Workbook()
    return wb.active


def test_none_returns_empty_string(ws) -> None:
    ws["A1"] = None
    assert cell_to_str(ws["A1"]) == ""


def test_empty_string_is_stripped(ws) -> None:
    ws["A1"] = "  hello  "
    assert cell_to_str(ws["A1"]) == "hello"


def test_int_becomes_decimal_string(ws) -> None:
    ws["A1"] = 8
    assert cell_to_str(ws["A1"]) == "8"


def test_int_zero(ws) -> None:
    ws["A1"] = 0
    assert cell_to_str(ws["A1"]) == "0"


def test_float_whole_number(ws) -> None:
    """8.0 should become '8' (common Excel width)."""
    ws["A1"] = 8.0
    assert cell_to_str(ws["A1"]) == "8"


def test_float_fractional(ws) -> None:
    ws["A1"] = 3.14
    assert cell_to_str(ws["A1"]) == "3.14"


def test_bool_true(ws) -> None:
    """Excel 0/1 sometimes auto-converts to bool — must produce '1'/'0'."""
    ws["A1"] = True
    assert cell_to_str(ws["A1"]) == "1"


def test_bool_false(ws) -> None:
    ws["A1"] = False
    assert cell_to_str(ws["A1"]) == "0"


def test_str_passes_through(ws) -> None:
    ws["A1"] = "DATA_WIDTH"
    assert cell_to_str(ws["A1"]) == "DATA_WIDTH"


def test_formula_raises(ws) -> None:
    """Cells with formulas must be detected BEFORE reading .value."""
    ws["A1"] = "=A2+1"
    with pytest.raises(FormulaCellError) as exc_info:
        cell_to_str(ws["A1"])
    assert exc_info.value.sheet == ws.title
    assert exc_info.value.row == 1
    assert exc_info.value.col == 1


def test_datetime_raises(ws) -> None:
    ws["A1"] = datetime.datetime(2026, 6, 1)
    with pytest.raises(UnsupportedCellTypeError) as exc_info:
        cell_to_str(ws["A1"])
    assert "datetime" in exc_info.value.type_name
    assert exc_info.value.row == 1


# ---- is_blank -------------------------------------------------------------

def test_is_blank_on_none(ws) -> None:
    ws["A1"] = None
    assert is_blank(ws["A1"])


def test_is_blank_on_empty_string(ws) -> None:
    ws["A1"] = ""
    assert is_blank(ws["A1"])


def test_is_blank_on_whitespace(ws) -> None:
    ws["A1"] = "   "
    assert is_blank(ws["A1"])


def test_not_blank_on_value(ws) -> None:
    ws["A1"] = "x"
    assert not is_blank(ws["A1"])


def test_not_blank_on_int(ws) -> None:
    ws["A1"] = 0
    # 0 is falsy in Python but still a meaningful cell
    assert not is_blank(ws["A1"])


def test_not_blank_on_formula(ws) -> None:
    """Formula cells are never blank (we don't trust them)."""
    ws["A1"] = "=1+1"
    assert not is_blank(ws["A1"])
