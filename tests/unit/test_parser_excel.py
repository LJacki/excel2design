"""Tests for parsers/excel.py (SPEC §2 + §3 integration)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excel2design.core.exceptions import (
    DuplicatePortError,
    HeaderMismatchError,
    MarkerMissingError,
    ModuleNotFoundError,
    PortValidationError,
)
from excel2design.core.models import (
    Direction,
    Parameter,
    ParamType,
    ResetType,
    SignalType,
)
from excel2design.parsers.excel import get_module, parse_array_dim, parse_workbook



def _write_sample_sheet(ws):
    """Write the canonical uart_rx sample data."""
    params = [
        ("DATA_WIDTH", 8, 32, "parameter", "数据位宽"),
        ("FIFO_DEPTH", 16, 32, "parameter", "FIFO 深度"),
        ("CLK_FREQ_MHZ", 100, 32, "parameter", "时钟频率(MHz)"),
    ]
    ports = [
        ("clk", "input", 1, "wire", "", "", "", 0, 0, "系统时钟"),
        ("rst_n", "input", 1, "wire", "", "", "", 0, 0, "异步低有效复位"),
        ("rx_pad", "input", 1, "wire", "", "", "", 0, 0, "串行输入"),
        ("baud_tick", "input", 1, "wire", "", "", "", 0, 0, "波特率 tick"),
        ("rx_data", "output", "DATA_WIDTH", "reg", "{DATA_WIDTH{1'b0}}", "clk", "async", 0, 0, "接收数据"),
        ("rx_valid", "output", 1, "reg", "1'b0", "clk", "async", 0, 0, "接收有效"),
        ("fifo_full", "output", 1, "reg", "1'b0", "clk", "async", 0, 0, "FIFO 满"),
        ("fifo_data", "output", "DATA_WIDTH", "reg", "{DATA_WIDTH{1'b0}}", "clk", "async", 1, 0, "FIFO 数据 (signed)"),
    ]
    ws.cell(1, 1, "# === PARAMETERS ===")
    ws.cell(2, 1, "name"); ws.cell(2, 2, "value"); ws.cell(2, 3, "width"); ws.cell(2, 4, "param_type"); ws.cell(2, 5, "comment")
    for i, row in enumerate(params):
        for j, val in enumerate(row):
            if val != "": ws.cell(3 + i, j + 1, val)
    nr = 3 + len(params)
    ws.cell(nr + 1, 1, "# === PORTS ===")
    ws.cell(nr + 2, 1, "name"); ws.cell(nr + 2, 2, "direction"); ws.cell(nr + 2, 3, "width")
    ws.cell(nr + 2, 4, "type"); ws.cell(nr + 2, 5, "default"); ws.cell(nr + 2, 6, "clock")
    ws.cell(nr + 2, 7, "reset_type"); ws.cell(nr + 2, 8, "signed"); ws.cell(nr + 2, 9, "interface"); ws.cell(nr + 2, 10, "comment")
    for i, row in enumerate(ports):
        for j, val in enumerate(row):
            if val != "": ws.cell(nr + 3 + i, j + 1, val)

# ---- Sample Excel fixture (generated on-the-fly) ---------------------------

@pytest.fixture(scope="module")
def sample_xlsx(tmp_path_factory):
    p = tmp_path_factory.mktemp("fixtures") / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "uart_rx"
    _write_sample_sheet(ws)
    wb.save(p)
    return p


def test_parse_sample_xlsx(sample_xlsx) -> None:
    """The canonical sample should parse cleanly."""
    modules = parse_workbook(sample_xlsx)
    assert len(modules) == 1
    m = modules[0]
    assert m.name == "uart_rx"


def test_sample_has_3_parameters(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    m = modules[0]
    assert len(m.parameters) == 3
    names = [p.name for p in m.parameters]
    assert names == ["DATA_WIDTH", "FIFO_DEPTH", "CLK_FREQ_MHZ"]
    assert m.parameters[0].value == "8"
    assert m.parameters[1].value == "16"


def test_sample_has_8_ports(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    m = modules[0]
    assert len(m.ports) == 8


def test_sample_inputs_outputs_classified(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    m = modules[0]
    assert len(m.inputs()) == 4
    assert len(m.outputs()) == 4
    assert len(m.inouts()) == 0


def test_sample_clk_is_wire_input(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    m = modules[0]
    clk = m.ports[0]
    assert clk.name == "clk"
    assert clk.direction == Direction.INPUT
    assert clk.type == SignalType.WIRE


def test_sample_rx_data_is_reg_output_param_width(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    m = modules[0]
    rx_data = next(p for p in m.ports if p.name == "rx_data")
    assert rx_data.direction == Direction.OUTPUT
    assert rx_data.type == SignalType.REG
    assert rx_data.width.is_parameter is True
    assert rx_data.width.raw == "DATA_WIDTH"
    assert rx_data.reset_type == ResetType.ASYNC
    assert rx_data.default == "{DATA_WIDTH{1'b0}}"


def test_sample_fifo_data_signed(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    m = modules[0]
    fifo_data = next(p for p in m.ports if p.name == "fifo_data")
    assert fifo_data.signed is True


def test_sample_preserves_excel_order(sample_xlsx) -> None:
    """Per SPEC §3.5.4, port order must match Excel row order."""
    modules = parse_workbook(sample_xlsx)
    m = modules[0]
    expected = [
        "clk", "rst_n", "rx_pad", "baud_tick",
        "rx_data", "rx_valid", "fifo_full", "fifo_data",
    ]
    assert [p.name for p in m.ports] == expected


def test_get_module_works(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    m = get_module(modules, "uart_rx")
    assert m.name == "uart_rx"


def test_get_module_missing_raises(sample_xlsx) -> None:
    modules = parse_workbook(sample_xlsx)
    with pytest.raises(ModuleNotFoundError) as exc_info:
        get_module(modules, "axi_crossbar")
    assert "axi_crossbar" in str(exc_info.value)
    assert "uart_rx" in str(exc_info.value)  # available list


# ---- File not found ------------------------------------------------------

def test_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        parse_workbook(tmp_path / "nonexistent.xlsx")


# ---- Helper to build in-memory Excel for failure cases -------------------

def _write_xlsx(path: Path, *, parameters: list[tuple] | None = None,
                ports: list[tuple] | None = None,
                include_param_marker: bool = True,
                include_port_marker: bool = True,
                param_header: list[str] | None = None,
                port_header: list[str] | None = None) -> None:
    """Build a minimal Excel file with the two-section layout.

    `parameters` and `ports` are tuples matching the canonical column order.
    """
    wb = Workbook()
    ws = wb.active
    row = 1

    if include_param_marker:
        ws.cell(row=row, column=1, value="# === PARAMETERS ===")
        row += 1
        ph = param_header or ["name", "value", "width", "param_type", "comment"]
        for c, h in enumerate(ph, 1):
            ws.cell(row=row, column=c, value=h)
        row += 1
        for p in parameters or []:
            for c, v in enumerate(p, 1):
                if v is not None:
                    ws.cell(row=row, column=c, value=v)
            row += 1
        row += 1  # blank

    if include_port_marker:
        ws.cell(row=row, column=1, value="# === PORTS ===")
        row += 1
        pth = port_header or [
            "name", "direction", "width", "type", "default",
            "clock", "reset_type", "signed", "interface", "comment",
        ]
        for c, h in enumerate(pth, 1):
            ws.cell(row=row, column=c, value=h)
        row += 1
        for p in ports or []:
            for c, v in enumerate(p, 1):
                if v is not None:
                    ws.cell(row=row, column=c, value=v)
            row += 1

    wb.save(path)


# ---- Error cases ----------------------------------------------------------

def test_missing_param_marker(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    _write_xlsx(p, include_param_marker=False, ports=[("clk", "input", "1", "wire")])
    with pytest.raises(MarkerMissingError) as exc_info:
        parse_workbook(p)
    assert "PARAMETERS" in exc_info.value.marker_name


def test_missing_port_marker(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    _write_xlsx(p, include_port_marker=False, parameters=[("W", "8")])
    with pytest.raises(MarkerMissingError) as exc_info:
        parse_workbook(p)
    assert "PORTS" in exc_info.value.marker_name


def test_header_mismatch(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    _write_xlsx(
        p,
        parameters=[("W", "8")],
        ports=[("clk", "input", "1", "wire")],
        port_header=["wrong", "header", "order", "x", "x", "x", "x", "x", "x", "x"],
    )
    with pytest.raises(HeaderMismatchError) as exc_info:
        parse_workbook(p)


def test_duplicate_port_name(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    _write_xlsx(
        p,
        parameters=[("W", "8")],
        ports=[
            ("clk", "input", "1", "wire", "", "", "", 0, 0, ""),
            ("clk", "input", "1", "wire", "", "", "", 0, 0, ""),  # duplicate
        ],
    )
    with pytest.raises(DuplicatePortError) as exc_info:
        parse_workbook(p)
    assert exc_info.value.name == "clk"
    assert len(exc_info.value.rows) == 2


def test_illegal_identifier_in_port(tmp_path: Path) -> None:
    """Reserved-word port name must raise with location."""
    p = tmp_path / "bad.xlsx"
    _write_xlsx(
        p,
        parameters=[("W", "8")],
        ports=[("reg", "input", "1", "wire", "", "", "", 0, 0, "")],
    )
    with pytest.raises(Exception) as exc_info:  # IdentifierError
        parse_workbook(p)
    assert "reg" in str(exc_info.value)


def test_illegal_identifier_in_sheet_name(tmp_path: Path) -> None:
    """Sheet name 'wire' is a reserved word → IdentifierError."""
    p = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "wire"  # reserved word
    ws.cell(row=1, column=1, value="# === PARAMETERS ===")
    ws.cell(row=2, column=1, value="name")
    ws.cell(row=2, column=2, value="value")
    ws.cell(row=3, column=1, value="W")
    ws.cell(row=3, column=2, value="8")
    ws.cell(row=4, column=1, value="# === PORTS ===")
    ws.cell(row=5, column=1, value="name")
    for c, h in enumerate(["direction", "width", "type", "default", "clock",
                           "reset_type", "signed", "interface", "comment"], 2):
        ws.cell(row=5, column=c, value=h)
    wb.save(p)
    with pytest.raises(Exception):  # IdentifierError on sheet name
        parse_workbook(p)


def test_unknown_width_parameter(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    _write_xlsx(
        p,
        parameters=[("W", "8")],
        ports=[("data", "output", "ADDR_WIDTH", "reg", "", "clk", "sync", 0, 0, "")],
    )
    with pytest.raises(Exception) as exc_info:  # UnknownParameterError
        parse_workbook(p)
    assert "ADDR_WIDTH" in str(exc_info.value)


def test_merged_cell_detected(tmp_path: Path) -> None:
    """We must reject merged cells (v0.3 doesn't support them).

    Build a sheet with a 2-cell horizontal merge inside the parameter section.
    """
    p = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "m"
    ws.cell(row=1, column=1, value="# === PARAMETERS ===")
    # header on row 2
    for c, h in enumerate(["name", "value", "width", "param_type", "comment"], 1):
        ws.cell(row=2, column=c, value=h)
    # Merge C3:D3 in the data area
    ws.cell(row=3, column=1, value="W")
    ws.cell(row=3, column=2, value="8")
    ws.cell(row=3, column=3, value="32")
    ws.merge_cells("C3:D3")
    # Ports section
    ws.cell(row=5, column=1, value="# === PORTS ===")
    for c, h in enumerate(["name", "direction", "width", "type", "default",
                           "clock", "reset_type", "signed", "interface", "comment"], 1):
        ws.cell(row=6, column=c, value=h)
    ws.cell(row=7, column=1, value="clk")
    ws.cell(row=7, column=2, value="input")
    ws.cell(row=7, column=3, value=1)
    ws.cell(row=7, column=4, value="wire")
    wb.save(p)
    from excel2design.core.exceptions import MergedCellError
    with pytest.raises(MergedCellError) as exc_info:
        parse_workbook(p)
    assert "C3" in exc_info.value.cell_range


def test_skip_blank_and_comment_rows(tmp_path: Path) -> None:
    """Blank lines and # comment lines must be skipped silently."""
    p = tmp_path / "ok.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "m"
    ws.cell(row=1, column=1, value="# === PARAMETERS ===")
    for c, h in enumerate(["name", "value", "width", "param_type", "comment"], 1):
        ws.cell(row=2, column=c, value=h)
    ws.cell(row=3, column=1, value="W")
    ws.cell(row=3, column=2, value=8)
    # blank row 4
    ws.cell(row=5, column=1, value="# a comment in params")
    # ports section
    ws.cell(row=6, column=1, value="# === PORTS ===")
    for c, h in enumerate(["name", "direction", "width", "type", "default",
                           "clock", "reset_type", "signed", "interface", "comment"], 1):
        ws.cell(row=7, column=c, value=h)
    # blank row 8
    ws.cell(row=9, column=1, value="# real port below")
    ws.cell(row=10, column=1, value="clk")
    ws.cell(row=10, column=2, value="input")
    ws.cell(row=10, column=3, value=1)
    ws.cell(row=10, column=4, value="wire")
    wb.save(p)
    modules = parse_workbook(p)
    assert len(modules) == 1
    assert [pp.name for pp in modules[0].ports] == ["clk"]
def test_inout_port(tmp_path: Path) -> None:
    p = tmp_path / "ok.xlsx"
    _write_xlsx(
        p,
        parameters=[("W", "8")],
        ports=[
            ("data", "inout", "W", "wire", "", "", "", 0, 0, "双向数据"),
        ],
    )
    modules = parse_workbook(p)
    m = modules[0]
    assert len(m.inouts()) == 1
    assert m.inouts()[0].name == "data"


def test_empty_port_section(tmp_path: Path) -> None:
    """A module with parameters but no ports is allowed."""
    p = tmp_path / "ok.xlsx"
    _write_xlsx(p, parameters=[("W", "8")], ports=[])
    modules = parse_workbook(p)
    assert len(modules) == 1
    assert len(modules[0].ports) == 0
    assert len(modules[0].parameters) == 1


def test_no_diff_on_repeat(sample_xlsx, tmp_path: Path) -> None:
    """Parsing the same file twice should give identical objects."""
    a = parse_workbook(sample_xlsx)
    b = parse_workbook(sample_xlsx)
    assert len(a) == len(b)
    assert [(m.name, len(m.ports), len(m.parameters)) for m in a] == \
           [(m.name, len(m.ports), len(m.parameters)) for m in b]


# ---- v0.6 Phase 14: parameter/port naming conflict warning ----------------

def test_param_port_collision_warning(tmp_path: Path) -> None:
    """Parameter and port with the same name (case-insensitive) → warning.

    Per SPEC §21 Phase 14: emit ``NamingConflictWarning`` (UserWarning
    subclass) listing the conflicting names. The module is still parsed
    successfully — this is a soft warning, not a hard error.
    """
    p = tmp_path / "collision.xlsx"
    _write_xlsx(
        p,
        parameters=[("WIDTH", "8")],
        ports=[
            ("clk", "input", "1", "wire", "", "", "", 0, 0, ""),
            ("width", "output", "WIDTH", "reg", "", "clk", "async", 0, 0, ""),
        ],
    )
    with pytest.warns(UserWarning) as record:
        modules = parse_workbook(p)
    assert len(modules) == 1
    # The warning should mention the parameter name AND the conflict type.
    msgs = [str(w.message) for w in record]
    assert any("WIDTH" in m and "重名" in m for m in msgs), (
        f"Expected a naming-conflict warning naming 'WIDTH', got: {msgs}"
    )


def test_param_port_collision_case_insensitive(tmp_path: Path) -> None:
    """Conflict is case-insensitive: ``Width`` (param) vs ``width`` (port)."""
    p = tmp_path / "ci.xlsx"
    _write_xlsx(
        p,
        parameters=[("Width", "8")],
        ports=[
            ("clk", "input", "1", "wire", "", "", "", 0, 0, ""),
            ("width", "output", "Width", "reg", "", "clk", "async", 0, 0, ""),
        ],
    )
    with pytest.warns(UserWarning):
        parse_workbook(p)


def test_no_conflict_no_warning(tmp_path: Path) -> None:
    """No name overlap → no NamingConflictWarning emitted."""
    p = tmp_path / "ok.xlsx"
    _write_xlsx(
        p,
        parameters=[("DATA_WIDTH", "8")],
        ports=[
            ("clk", "input", "1", "wire", "", "", "", 0, 0, ""),
            ("data", "output", "DATA_WIDTH", "reg", "", "clk", "async", 0, 0, ""),
        ],
    )
    import warnings as _w
    with _w.catch_warnings(record=True) as captured:
        _w.simplefilter("always")
        parse_workbook(p)
    # Should not have raised any UserWarning about naming conflict.
    conflict_warnings = [
        w for w in captured
        if "naming" in str(w.message).lower() or "重名" in str(w.message)
    ]
    assert conflict_warnings == [], (
        f"Expected no naming conflict warning, got: "
        f"{[str(w.message) for w in conflict_warnings]}"
    )


# ---- v0.6 Phase 12.2: array_dim parsing ----------------------------------

def test_parse_array_dim_empty_returns_none() -> None:
    assert parse_array_dim("") is None
    assert parse_array_dim(None) is None
    assert parse_array_dim("   ") is None


def test_parse_array_dim_single_dim() -> None:
    assert parse_array_dim("[7:0]") == [(7, 0)]
    assert parse_array_dim("[0:0]") == [(0, 0)]


def test_parse_array_dim_multi_dim() -> None:
    assert parse_array_dim("[3:0][1:0]") == [(3, 0), (1, 0)]
    assert parse_array_dim("[7:0][3:0][1:0]") == [(7, 0), (3, 0), (1, 0)]


def test_parse_array_dim_whitespace_tolerated() -> None:
    assert parse_array_dim("  [7:0]  ") == [(7, 0)]
    assert parse_array_dim(" [3:0] [1:0] ") == [(3, 0), (1, 0)]


def test_parse_array_dim_invalid_raises() -> None:
    with pytest.raises(PortValidationError):
        parse_array_dim("[7:0] foo")
    with pytest.raises(PortValidationError):
        parse_array_dim("foo")
    with pytest.raises(PortValidationError):
        parse_array_dim("[7:0] [3")
    with pytest.raises(PortValidationError):
        parse_array_dim("[7:0][a:b]")


def test_parse_array_dim_reversed_range_raises() -> None:
    """hi < lo → PortValidationError (SPEC §21: array_dim 范围必须 hi >= lo)."""
    with pytest.raises(PortValidationError):
        parse_array_dim("[0:7]")
    with pytest.raises(PortValidationError):
        parse_array_dim("[3:5]")


# ---- v0.6 Phase 12.2: Excel-level array_dim round-trip --------------------

def _write_xlsx_with_array_dim(path: Path, ports: list[tuple]) -> None:
    """Build an Excel with the canonical 10-col port header plus optional
    11th `array_dim` column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "m"
    ws.cell(row=1, column=1, value="# === PARAMETERS ===")
    for c, h in enumerate(["name", "value", "width", "param_type", "comment"], 1):
        ws.cell(row=2, column=c, value=h)
    ws.cell(row=3, column=1, value="W"); ws.cell(row=3, column=2, value="8")
    ws.cell(row=5, column=1, value="# === PORTS ===")
    headers = ["name", "direction", "width", "type", "default",
               "clock", "reset_type", "signed", "interface", "comment",
               "array_dim"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=6, column=c, value=h)
    for i, row_vals in enumerate(ports):
        for c, v in enumerate(row_vals, 1):
            if v is not None and v != "":
                ws.cell(row=7 + i, column=c, value=v)
    wb.save(path)


def test_excel_array_dim_column_parses(tmp_path: Path) -> None:
    """An Excel with the 11th `array_dim` column should populate Port.array_dim."""
    p = tmp_path / "arr.xlsx"
    _write_xlsx_with_array_dim(p, ports=[
        ("clk", "input", 1, "wire", "", "", "", 0, 0, "系统时钟", ""),
        ("array_port", "output", "W", "reg", "{W{1'b0}}", "clk", "sync", 0, 0,
         "端口数组", "[7:0]"),
    ])
    modules = parse_workbook(p)
    m = modules[0]
    clk = m.ports[0]
    assert clk.array_dim is None  # empty cell → None
    arr = m.ports[1]
    assert arr.array_dim == [(7, 0)]


def test_excel_array_dim_invalid_raises(tmp_path: Path) -> None:
    """Bad array_dim format → PortValidationError with sheet/row info."""
    p = tmp_path / "bad.xlsx"
    _write_xlsx_with_array_dim(p, ports=[
        ("clk", "input", 1, "wire", "", "", "", 0, 0, "系统时钟", ""),
        ("bad_port", "output", "W", "reg", "", "clk", "sync", 0, 0,
         "错误格式", "[7:0]foo"),
    ])
    with pytest.raises(PortValidationError) as exc_info:
        parse_workbook(p)
    assert "array_dim" in str(exc_info.value)


def test_excel_no_array_dim_column_still_parses(tmp_path: Path) -> None:
    """Old 10-column fixtures (no array_dim header) must remain parseable."""
    # Reuse the existing helper from above (_write_xlsx)
    p = tmp_path / "old.xlsx"
    _write_xlsx(
        p,
        parameters=[("W", "8")],
        ports=[("clk", "input", "1", "wire", "", "", "", 0, 0, "时钟")],
    )
    modules = parse_workbook(p)
    assert modules[0].ports[0].array_dim is None
